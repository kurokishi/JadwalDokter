"""
Generate formatted Excel output
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import io
from app.config import AppConfig


class ExcelGenerator:
    """Generate formatted Excel file"""
    
    def __init__(self):
        self.config = AppConfig()
        
        # Define styles
        self.header_fill = PatternFill(
            start_color=self.config.COLOR_HEADER,
            end_color=self.config.COLOR_HEADER,
            fill_type="solid"
        )
        
        self.reguler_fill = PatternFill(
            start_color=self.config.COLOR_REGULER,
            end_color=self.config.COLOR_REGULER,
            fill_type="solid"
        )
        
        self.eksekutif_fill = PatternFill(
            start_color=self.config.COLOR_EKSEKUTIF,
            end_color=self.config.COLOR_EKSEKUTIF,
            fill_type="solid"
        )
        
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.normal_font = Font(size=10)
        self.bold_font = Font(bold=True, size=10)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def generate_excel(self, df: pd.DataFrame, sheet_name: str = "Jadwal") -> bytes:
        """
        Generate formatted Excel file from DataFrame
        
        Args:
            df: DataFrame with grid format
            sheet_name: Name of the sheet
            
        Returns:
            Excel file as bytes
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        headers = list(df.columns)
        self._write_headers(ws, headers)
        
        # Write data rows
        self._write_data_rows(ws, df, headers)
        
        # Apply formatting
        self._apply_formatting(ws, len(df), len(headers))
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _write_headers(self, ws, headers):
        """Write and format headers"""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = self.center_alignment
    
    def _write_data_rows(self, ws, df, headers):
        """Write data rows with formatting"""
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, col_name in enumerate(headers, 1):
                value = row_data[col_name]
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                
                # Apply formatting based on column and value
                self._apply_cell_formatting(cell, col_idx, col_name, value, headers)
    
    def _apply_cell_formatting(self, cell, col_idx, col_name, value, headers):
        """Apply formatting to individual cell"""
        # First 5 columns (POLI, JENIS, HARI, DOKTER, JAM)
        if col_idx <= 5:
            cell.alignment = self.left_alignment
            cell.font = self.normal_font
            
            # Bold for DOKTER column
            if col_name == 'DOKTER':
                cell.font = self.bold_font
        else:
            # Time slot columns
            cell.alignment = self.center_alignment
            
            if value == 'R':
                cell.fill = self.reguler_fill
                cell.font = self.bold_font
            elif value == 'E':
                cell.fill = self.eksekutif_fill
                cell.font = self.bold_font
            else:
                cell.font = self.normal_font
    
    def _apply_formatting(self, ws, num_rows, num_cols):
        """Apply overall formatting to worksheet"""
        # Auto-adjust column widths
        for col in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, num_rows + 2):  # +2 for header and 0-index
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze panes (freeze header row and first 5 columns)
        ws.freeze_panes = ws['F2']
        
        # Add filter to header row
        ws.auto_filter.ref = ws.dimensions
    
    def generate_simple_excel(self, df: pd.DataFrame) -> bytes:
        """Generate simple Excel without formatting (faster)"""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Jadwal')
        output.seek(0)
        return output.getvalue()
