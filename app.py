import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, time
import openpyxl
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

# Konfigurasi halaman
st.set_page_config(
    page_title="Jadwal Dokter - RS",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Cache data untuk performa
@st.cache_data
def load_data(file_path="jadwal hafis.xlsx"):
    """Memuat data dari file Excel"""
    try:
        # Baca file Excel
        df_raw = pd.read_excel(file_path, sheet_name=0)
        
        # Proses data mentah
        processed_data = process_excel_data(df_raw)
        return processed_data
    except FileNotFoundError:
        st.error(f"File {file_path} tidak ditemukan. Membuat template kosong.")
        return create_empty_template()
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return create_empty_template()

@st.cache_data
def load_result_data(file_path="jadwal_hasil_2026.xlsx"):
    """Memuat data hasil dari file Excel kedua"""
    try:
        df_result = pd.read_excel(file_path, sheet_name=0)
        return df_result
    except:
        return None

def process_excel_data(df_raw):
    """Memproses data mentah dari Excel menjadi format yang mudah digunakan"""
    # Reset index dan rename kolom
    df = df_raw.copy()
    
    # Identifikasi kolom hari
    day_columns = ['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT', 'SABTU']
    
    # Buat list untuk menyimpan data terstruktur
    data_list = []
    
    current_ksm = None
    current_dokter = None
    
    for idx, row in df.iterrows():
        # Periksa apakah baris berisi KSM
        if pd.notna(row['KSM']):
            current_ksm = row['KSM']
        
        # Periksa apakah baris berisi nama dokter
        if pd.notna(row['Nama dokter spesialis/ sub spesialis']):
            current_dokter = row['Nama dokter spesialis/ sub spesialis']
        
        # Periksa tipe jadwal (JAM KERJA, REGULER, EKSEKUTIF)
        if pd.notna(row['POLI']):
            jadwal_type = row['POLI']
            
            # Proses setiap hari
            for day in day_columns:
                if day in df.columns:
                    jam_value = row[day]
                    if pd.notna(jam_value) and str(jam_value).strip() not in ['', '-', 'nan']:
                        data_list.append({
                            'KSM': current_ksm,
                            'Dokter': current_dokter,
                            'Jenis': jadwal_type,
                            'Hari': day,
                            'Jam': str(jam_value)
                        })
    
    return pd.DataFrame(data_list)

def create_empty_template():
    """Membuat template kosong jika file tidak ada"""
    return pd.DataFrame(columns=['KSM', 'Dokter', 'Jenis', 'Hari', 'Jam'])

def parse_time_slots(jam_string):
    """Mengurai string jam menjadi slot waktu 30 menit"""
    if not jam_string or str(jam_string).strip() in ['', '-', 'nan', 'None']:
        return []
    
    slots = []
    jam_string = str(jam_string).replace('.', ':')
    
    # Handle multiple time ranges
    time_ranges = jam_string.split(',')
    
    for time_range in time_ranges:
        time_range = time_range.strip()
        
        # Skip jika kosong
        if not time_range:
            continue
        
        # Coba parsing format berbeda
        try:
            if '-' in time_range:
                start_str, end_str = time_range.split('-')
                start_time = parse_time(start_str.strip())
                end_time = parse_time(end_str.strip())
                
                # Generate slots setiap 30 menit
                current = start_time
                while current < end_time:
                    slots.append(current.strftime('%H:%M'))
                    # Tambah 30 menit
                    current = datetime.combine(datetime.today(), current)
                    current = (current + pd.Timedelta(minutes=30)).time()
            else:
                # Single time
                single_time = parse_time(time_range)
                slots.append(single_time.strftime('%H:%M'))
        except:
            # Jika parsing gagal, skip
            continue
    
    return slots

def parse_time(time_str):
    """Mengonversi string waktu ke time object"""
    time_str = str(time_str).strip().replace('.', ':')
    
    # Coba berbagai format
    formats = ['%H:%M', '%H.%M', '%H:%M:%S', '%H.%M.%S']
    
    for fmt in formats:
        try:
            return datetime.strptime(time_str, fmt).time()
        except:
            continue
    
    # Default fallback
    return datetime.strptime("00:00", "%H:%M").time()

def create_schedule_grid(df, show_saturday=False):
    """Membuat grid jadwal seperti Excel"""
    # Filter hari
    days = ['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT']
    if show_saturday:
        days.append('SABTU')
    
    # Buat slot waktu
    time_slots = []
    for hour in range(7, 15):  # 07:00 sampai 14:30
        for minute in [0, 30]:
            time_slots.append(f"{hour:02d}:{minute:02d}")
    
    # Buat DataFrame untuk grid
    grid_data = []
    
    for _, row in df.iterrows():
        if row['Hari'] in days:
            time_slots_list = parse_time_slots(row['Jam'])
            
            for time_slot in time_slots_list:
                grid_data.append({
                    'KSM': row['KSM'],
                    'Dokter': row['Dokter'],
                    'Hari': row['Hari'],
                    'Jenis': row['Jenis'],
                    'Waktu': time_slot,
                    'Slot': 'R' if 'REGULER' in row['Jenis'] else 'E' if 'EKSEKUTIF' in row['Jenis'] else 'K'
                })
    
    grid_df = pd.DataFrame(grid_data)
    
    if grid_df.empty:
        return pd.DataFrame()
    
    # Buat pivot table
    pivot_df = grid_df.pivot_table(
        index=['KSM', 'Dokter', 'Hari'],
        columns='Waktu',
        values='Slot',
        aggfunc='first'
    ).fillna('')
    
    # Urutkan kolom waktu
    pivot_df = pivot_df.reindex(columns=time_slots, fill_value='')
    
    return pivot_df.reset_index()

def create_time_slot_chart(df, show_saturday=False):
    """Membuat chart visualisasi ketersediaan slot"""
    days = ['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT']
    if show_saturday:
        days.append('SABTU')
    
    # Filter data
    filtered_df = df[df['Hari'].isin(days)].copy()
    
    if filtered_df.empty:
        return None
    
    # Hitung jumlah slot per hari per jenis
    chart_data = []
    
    for day in days:
        day_df = filtered_df[filtered_df['Hari'] == day]
        
        reguler_count = len(day_df[day_df['Jenis'] == 'REGULER'])
        eksekutif_count = len(day_df[day_df['Jenis'] == 'EKSEKUTIF'])
        
        chart_data.append({
            'Hari': day,
            'Reguler': reguler_count,
            'Eksekutif': eksekutif_count
        })
    
    chart_df = pd.DataFrame(chart_data)
    
    # Buat stacked bar chart
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=chart_df['Hari'],
        y=chart_df['Reguler'],
        name='Reguler',
        marker_color='green',
        text=chart_df['Reguler'],
        textposition='auto'
    ))
    
    fig.add_trace(go.Bar(
        x=chart_df['Hari'],
        y=chart_df['Eksekutif'],
        name='Eksekutif',
        marker_color='blue',
        text=chart_df['Eksekutif'],
        textposition='auto'
    ))
    
    fig.update_layout(
        title='Distribusi Jadwal per Hari',
        xaxis_title='Hari',
        yaxis_title='Jumlah Jadwal',
        barmode='stack',
        height=400
    )
    
    return fig

def validate_executive_slots(schedule_data, new_slot, max_slots=7):
    """Validasi slot eksekutif tidak melebihi batas"""
    # Ekstrak info dari new_slot
    hari = new_slot['Hari']
    waktu = new_slot['Waktu']
    
    # Filter slot yang sama
    same_slots = schedule_data[
        (schedule_data['Hari'] == hari) &
        (schedule_data['Waktu'] == waktu) &
        (schedule_data['Jenis'] == 'EKSEKUTIF')
    ]
    
    # Hitung dokter yang sudah ada (exclude dokter yang sama jika edit)
    existing_count = len(same_slots)
    
    if existing_count >= max_slots:
        return False, f"Slot eksekutif pada {hari} {waktu} sudah penuh (maks {max_slots} dokter)"
    
    return True, "Valid"

def save_to_excel(df, file_path="jadwal_hafis_baru.xlsx"):
    """Menyimpan data ke Excel"""
    try:
        # Buat workbook baru
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Header sesuai format asli
        headers = ['KSM', 'Nama dokter spesialis/ sub spesialis', 'POLI', 
                  'SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT', 'SABTU']
        
        # Tulis header
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Tulis data
        row_idx = 2
        current_ksm = None
        current_dokter = None
        
        # Group data
        grouped = df.groupby(['KSM', 'Dokter', 'Jenis', 'Hari'])
        
        # Buat struktur untuk menulis
        write_data = {}
        
        for (ksm, dokter, jenis, hari), group in grouped:
            if (ksm, dokter) not in write_data:
                write_data[(ksm, dokter)] = {}
            
            if jenis not in write_data[(ksm, dokter)]:
                write_data[(ksm, dokter)][jenis] = {}
            
            # Gabungkan jam untuk hari yang sama
            jam_values = group['Jam'].tolist()
            # Gabungkan jika multiple
            if len(jam_values) > 1:
                jam_str = ', '.join(jam_values)
            else:
                jam_str = jam_values[0] if jam_values else ''
            
            write_data[(ksm, dokter)][jenis][hari] = jam_str
        
        # Tulis ke Excel
        for (ksm, dokter), jenis_data in write_data.items():
            # Baris KSM dan Dokter
            if current_ksm != ksm:
                ws.cell(row=row_idx, column=1, value=ksm)
                ws.cell(row=row_idx, column=2, value=dokter)
                ws.cell(row=row_idx, column=3, value='JAM KERJA')
                current_ksm = ksm
                current_dokter = dokter
                row_idx += 1
            elif current_dokter != dokter:
                ws.cell(row=row_idx, column=2, value=dokter)
                ws.cell(row=row_idx, column=3, value='JAM KERJA')
                current_dokter = dokter
                row_idx += 1
            
            # Baris REGULER
            ws.cell(row=row_idx, column=3, value='REGULER')
            if 'REGULER' in jenis_data:
                for day_idx, day in enumerate(['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT', 'SABTU'], 4):
                    if day in jenis_data['REGULER']:
                        ws.cell(row=row_idx, column=day_idx, value=jenis_data['REGULER'][day])
            row_idx += 1
            
            # Baris EKSEKUTIF
            ws.cell(row=row_idx, column=3, value='EKSEKUTIF')
            if 'EKSEKUTIF' in jenis_data:
                for day_idx, day in enumerate(['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT', 'SABTU'], 4):
                    if day in jenis_data['EKSEKUTIF']:
                        ws.cell(row=row_idx, column=day_idx, value=jenis_data['EKSEKUTIF'][day])
            row_idx += 1
            
            # Tambah baris kosong antar dokter
            row_idx += 1
        
        # Simpan ke buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    except Exception as e:
        st.error(f"Error saving to Excel: {str(e)}")
        return None

def main():
    # Judul aplikasi
    st.title("🏥 Sistem Penjadwalan Dokter")
    st.markdown("---")
    
    # Sidebar untuk filter
    with st.sidebar:
        st.header("⚙️ Filter & Kontrol")
        
        # Upload file
        uploaded_file = st.file_uploader("Unggah File Jadwal", type=['xlsx', 'xls'])
        
        if uploaded_file is not None:
            # Simpan file sementara
            with open("jadwal_hafis.xlsx", "wb") as f:
                f.write(uploaded_file.getbuffer())
            st.success("File berhasil diunggah!")
        
        # Pilihan tampilan
        st.subheader("Tampilan")
        view_mode = st.radio(
            "Mode Tampilan",
            ["View", "Edit"],
            horizontal=True
        )
        
        # Filter Poli/KSM
        data = load_data()
        poli_options = ['Semua'] + sorted(data['KSM'].dropna().unique().tolist())
        selected_poli = st.selectbox("Filter Poli/KSM", poli_options)
        
        # Filter Hari
        st.subheader("Filter Hari")
        show_saturday = st.checkbox("Tampilkan Sabtu", value=False)
        
        hari_options = ['Semua', 'SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT']
        if show_saturday:
            hari_options.append('SABTU')
        selected_hari = st.multiselect(
            "Pilih Hari",
            hari_options,
            default=['Semua']
        )
        
        # Filter Jenis
        st.subheader("Filter Jenis Jadwal")
        show_reguler = st.checkbox("Reguler", value=True)
        show_eksekutif = st.checkbox("Eksekutif", value=True)
        
        # Search dokter
        dokter_search = st.text_input("🔍 Cari Dokter")
        
        # Statistik
        st.markdown("---")
        st.subheader("📊 Statistik")
        
        if not data.empty:
            total_dokter = data['Dokter'].nunique()
            total_jadwal = len(data)
            
            st.metric("Total Dokter", total_dokter)
            st.metric("Total Jadwal", total_jadwal)
            
            # Load data hasil untuk perbandingan
            result_data = load_result_data()
            if result_data is not None:
                st.metric("Jadwal Tervalidasi", len(result_data))
    
    # Main content area
    if view_mode == "View":
        display_view_mode(data, selected_poli, selected_hari, show_saturday, 
                         show_reguler, show_eksekutif, dokter_search)
    else:
        display_edit_mode(data)

def display_view_mode(data, selected_poli, selected_hari, show_saturday, 
                     show_reguler, show_eksekutif, dokter_search):
    """Menampilkan mode view"""
    
    # Filter data berdasarkan pilihan
    filtered_data = data.copy()
    
    # Filter Poli
    if selected_poli != 'Semua':
        filtered_data = filtered_data[filtered_data['KSM'] == selected_poli]
    
    # Filter Hari
    if 'Semua' not in selected_hari and selected_hari:
        filtered_data = filtered_data[filtered_data['Hari'].isin(selected_hari)]
    
    # Filter Jenis
    jenis_filter = []
    if show_reguler:
        jenis_filter.append('REGULER')
    if show_eksekutif:
        jenis_filter.append('EKSEKUTIF')
    
    if jenis_filter:
        filtered_data = filtered_data[filtered_data['Jenis'].isin(jenis_filter)]
    
    # Filter search dokter
    if dokter_search:
        filtered_data = filtered_data[
            filtered_data['Dokter'].str.contains(dokter_search, case=False, na=False)
        ]
    
    # Tampilkan data dalam dua tab
    tab1, tab2, tab3 = st.tabs(["📋 Tabel Data", "📅 Grid Jadwal", "📊 Visualisasi"])
    
    with tab1:
        if not filtered_data.empty:
            st.dataframe(
                filtered_data,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "KSM": st.column_config.TextColumn("Poli/KSM"),
                    "Dokter": st.column_config.TextColumn("Nama Dokter"),
                    "Jenis": st.column_config.TextColumn("Jenis Jadwal"),
                    "Hari": st.column_config.TextColumn("Hari"),
                    "Jam": st.column_config.TextColumn("Jam Praktik")
                }
            )
            
            # Tombol download
            csv = filtered_data.to_csv(index=False)
            st.download_button(
                label="📥 Download Data (CSV)",
                data=csv,
                file_name="jadwal_dokter.csv",
                mime="text/csv"
            )
        else:
            st.info("Tidak ada data yang sesuai dengan filter.")
    
    with tab2:
        st.subheader("Grid Jadwal (Excel-like View)")
        
        # Buat grid jadwal
        grid_df = create_schedule_grid(filtered_data, show_saturday)
        
        if not grid_df.empty:
            # Tampilkan grid dengan styling
            styled_df = grid_df.style.applymap(
                lambda x: 'background-color: #d4edda; color: #155724; font-weight: bold;' 
                if x == 'R' else 
                'background-color: #cce5ff; color: #004085; font-weight: bold;' 
                if x == 'E' else 
                'background-color: #f8f9fa; color: #6c757d;'
                if x == 'K' else 
                'background-color: #e9ecef;'
            )
            
            st.dataframe(
                styled_df,
                use_container_width=True,
                height=600
            )
        else:
            st.info("Tidak ada jadwal untuk ditampilkan dalam format grid.")
    
    with tab3:
        st.subheader("Visualisasi Ketersediaan")
        
        # Chart distribusi
        fig = create_time_slot_chart(filtered_data, show_saturday)
        if fig:
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Tidak cukup data untuk visualisasi.")
        
        # Heatmap ketersediaan
        if not filtered_data.empty:
            st.subheader("Heatmap Ketersediaan per Waktu")
            
            # Buat heatmap data
            heatmap_data = []
            days = ['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT']
            if show_saturday:
                days.append('SABTU')
            
            for day in days:
                day_data = {'Hari': day}
                for hour in range(7, 15):
                    for minute in [0, 30]:
                        time_slot = f"{hour:02d}:{minute:02d}"
                        
                        # Hitung jumlah jadwal di slot ini
                        count = len([
                            row for _, row in filtered_data.iterrows()
                            if row['Hari'] == day and 
                            time_slot in parse_time_slots(row['Jam'])
                        ])
                        
                        day_data[time_slot] = count
                
                heatmap_data.append(day_data)
            
            heatmap_df = pd.DataFrame(heatmap_data)
            heatmap_df = heatmap_df.set_index('Hari')
            
            # Tampilkan heatmap
            fig_heatmap = px.imshow(
                heatmap_df,
                labels=dict(x="Waktu", y="Hari", color="Jumlah Jadwal"),
                aspect="auto"
            )
            
            fig_heatmap.update_layout(height=400)
            st.plotly_chart(fig_heatmap, use_container_width=True)

def display_edit_mode(data):
    """Menampilkan mode edit"""
    
    st.header("✏️ Edit Jadwal Dokter")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        # Form untuk menambah/edit jadwal
        with st.form("edit_jadwal_form"):
            st.subheader("Tambah/Edit Jadwal")
            
            # Pilih aksi
            action = st.radio(
                "Aksi",
                ["Tambah Jadwal Baru", "Edit Jadwal Existing"],
                horizontal=True
            )
            
            # Pilih dokter
            dokter_list = sorted(data['Dokter'].dropna().unique().tolist())
            selected_dokter = st.selectbox("Pilih Dokter", dokter_list)
            
            # Pilih Poli
            poli_list = sorted(data['KSM'].dropna().unique().tolist())
            selected_poli = st.selectbox("Pilih Poli/KSM", poli_list)
            
            # Pilih Hari
            hari_options = ['SENIN', 'SELASA', 'RABU', 'KAMIS', 'JUMAT', 'SABTU']
            selected_hari = st.selectbox("Pilih Hari", hari_options)
            
            # Pilih Jenis
            jenis_options = ['REGULER', 'EKSEKUTIF', 'JAM KERJA']
            selected_jenis = st.selectbox("Pilih Jenis Jadwal", jenis_options)
            
            # Input jam
            st.subheader("Input Jam Praktik")
            col_time1, col_time2 = st.columns(2)
            
            with col_time1:
                start_time = st.time_input("Jam Mulai", value=time(7, 30))
            
            with col_time2:
                end_time = st.time_input("Jam Selesai", value=time(14, 0))
            
            # Multiple time slots
            st.write("Atau input manual (format: 07.30-14.00, 08.00-09.00)")
            manual_time = st.text_input("Jam Manual")
            
            # Tombol submit
            submitted = st.form_submit_button("💾 Simpan Jadwal")
            
            if submitted:
                # Validasi
                if start_time >= end_time:
                    st.error("Jam mulai harus sebelum jam selesai!")
                else:
                    # Format jam
                    if manual_time:
                        jam_input = manual_time
                    else:
                        jam_input = f"{start_time.strftime('%H.%M')}-{end_time.strftime('%H.%M')}"
                    
                    # Buat entry baru
                    new_entry = {
                        'KSM': selected_poli,
                        'Dokter': selected_dokter,
                        'Jenis': selected_jenis,
                        'Hari': selected_hari,
                        'Jam': jam_input
                    }
                    
                    # Validasi khusus untuk eksekutif
                    if selected_jenis == 'EKSEKUTIF':
                        is_valid, message = validate_executive_slots(
                            data, 
                            {'Hari': selected_hari, 'Waktu': start_time.strftime('%H:%M')}
                        )
                        
                        if not is_valid:
                            st.warning(message)
                        else:
                            # Tambah ke data
                            data = pd.concat([data, pd.DataFrame([new_entry])], ignore_index=True)
                            st.success("Jadwal berhasil ditambahkan!")
                    else:
                        # Tambah ke data
                        data = pd.concat([data, pd.DataFrame([new_entry])], ignore_index=True)
                        st.success("Jadwal berhasil ditambahkan!")
    
    with col2:
        st.subheader("🗑️ Hapus Jadwal")
        
        if not data.empty:
            # Pilih jadwal untuk dihapus
            jadwal_list = []
            for idx, row in data.iterrows():
                jadwal_list.append(f"{row['Dokter']} - {row['Hari']} - {row['Jenis']}: {row['Jam']}")
            
            selected_jadwal_to_delete = st.selectbox(
                "Pilih Jadwal untuk Dihapus",
                jadwal_list
            )
            
            if st.button("🗑️ Hapus Jadwal Terpilih", type="secondary"):
                # Temukan index jadwal yang dipilih
                if selected_jadwal_to_delete:
                    # Parse selection
                    parts = selected_jadwal_to_delete.split(" - ")
                    if len(parts) >= 3:
                        dokter_name = parts[0]
                        hari_name = parts[1]
                        jenis_jam = parts[2]
                        
                        # Split jenis dan jam
                        if ": " in jenis_jam:
                            jenis_part, jam_part = jenis_jam.split(": ", 1)
                        else:
                            jenis_part = jenis_jam
                            jam_part = ""
                        
                        # Hapus dari data
                        mask = (
                            (data['Dokter'] == dokter_name) &
                            (data['Hari'] == hari_name) &
                            (data['Jenis'] == jenis_part.strip())
                        )
                        
                        if jam_part:
                            mask = mask & (data['Jam'] == jam_part.strip())
                        
                        data = data[~mask].reset_index(drop=True)
                        st.success("Jadwal berhasil dihapus!")
        
        st.markdown("---")
        st.subheader("💾 Simpan Perubahan")
        
        if st.button("💾 Simpan ke Excel", type="primary"):
            excel_buffer = save_to_excel(data)
            if excel_buffer:
                st.download_button(
                    label="📥 Download File Excel",
                    data=excel_buffer,
                    file_name="jadwal_dokter_updated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
    
    # Tampilkan data saat ini
    st.markdown("---")
    st.subheader("📋 Data Saat Ini")
    
    if not data.empty:
        st.dataframe(
            data,
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("Belum ada data jadwal.")

if __name__ == "__main__":
    main()
